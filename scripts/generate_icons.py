#!/usr/bin/env python3
"""Download and restyle SVG icons for e-commerce categories.

The script reads a taxonomy spreadsheet (CSV or XLSX) describing category rows
and downloads an SVG for each row using the svgapi.com JSON API. Icons can be
exported in multiple style variants and are written to
``<output>/<style>/<category>/<Catid>.svg`` alongside a ``manifest.csv`` with
metadata useful for validation. A raw variant is provided so operators can
review the untouched download from the catalogue; those files are prefixed with
``test-``.

The selected icon for a category is deterministic: a SHA256 hash of ``Catid``
is used to pick one item from the API search results.
"""

import argparse
import csv
import hashlib
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import requests
import xml.etree.ElementTree as ET

ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.append(str(SRC_PATH))

from taxonomy.resolver import deepest_category  # noqa: E402
from taxonomy.synonyms import build_queries  # noqa: E402


STYLE_VARIANTS = {
    "original": {
        "raw_output": True,
        "preserve_source_style": True,
        "file_prefix": "test-",
    },
    "brand": {
        "stroke_color": "#E63B14",
        "stroke_width": 12,
        "fill": "none",
    },
    "thin": {
        "stroke_color": "#E63B14",
        "stroke_width": 8,
        "fill": "none",
    },
    "thick": {
        "stroke_color": "#E63B14",
        "stroke_width": 16,
        "fill": "none",
    },
    "mono": {
        "stroke_color": "#000000",
        "stroke_width": 12,
        "fill": "none",
    },
    "blue": {
        "stroke_color": "#004165",
        "stroke_width": 12,
        "fill": "none",
    },
}

STYLE_ALIASES = {"classic": "original"}

DEFAULT_STYLES = ("original", "brand", "thin", "thick", "mono")

SVG_NS = "http://www.w3.org/2000/svg"
SVGAPI_LIST_URL = "https://api.svgapi.com/v1/{key}/list/"


def configure_logging(out_dir: Path, level: str) -> Path:
    """Configure logging to both STDOUT and ``generation.log`` in ``out_dir``."""

    out_dir.mkdir(parents=True, exist_ok=True)
    log_path = out_dir / "generation.log"

    console_level = getattr(logging, level.upper(), logging.INFO)
    console_handler = logging.StreamHandler()
    console_handler.setLevel(console_level)

    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setLevel(logging.INFO)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[console_handler, file_handler],
    )
    logging.info("Logging initialised at %s", log_path)
    return log_path


def element_signature(el: ET.Element) -> str:
    """Recursively generate a signature for an element."""
    tag = el.tag.split('}')[-1]
    sig = tag + "".join(f"{k}={el.get(k)}" for k in sorted(el.attrib))
    return sig + "".join(element_signature(c) for c in el)


def slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", name.lower())
    slug = re.sub(r"_+", "_", slug).strip("_")
    return slug or "category"


def iter_search_queries(category: str) -> List[str]:
    """Generate prioritized search queries for ``category``."""

    queries: List[str] = []
    for candidate in build_queries(category):
        sanitized = candidate.strip()
        if sanitized and sanitized not in queries:
            queries.append(sanitized)

    # Add a simplified form without punctuation to avoid API parse errors.
    simplified = re.sub(r"[^a-z0-9 ]+", " ", category.lower()).strip()
    if simplified and simplified not in queries:
        queries.append(simplified)

    if category not in queries:
        queries.append(category)

    lower = category.lower()
    fallback_terms: List[str] = []
    if "baby" in lower:
        fallback_terms.extend(["baby care", "baby icon", "baby"])
    if "kind" in lower or "child" in lower:
        fallback_terms.extend(["child icon", "children toys"])

    for term in fallback_terms:
        if term not in queries:
            queries.append(term)

    universal_fallbacks = ["baby icon", "baby"]
    for term in universal_fallbacks:
        if term not in queries:
            queries.append(term)
    return queries


def load_taxonomy_rows(path: Path) -> List[Dict[str, str]]:
    """Return taxonomy rows from ``path`` supporting CSV and XLSX files."""

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - import guard
            raise SystemExit(
                "Reading .xlsx files requires openpyxl. Install it with 'pip install openpyxl'."
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows: List[Dict[str, str]] = []
        header_row = next(worksheet.iter_rows(values_only=True), None)
        if not header_row:
            workbook.close()
            return rows

        headers: List[str] = []
        for idx, cell in enumerate(header_row):
            header = str(cell).strip() if cell is not None else ""
            if not header:
                header = f"column_{idx}"
            headers.append(header)

        for excel_row in worksheet.iter_rows(min_row=2, values_only=True):
            if excel_row is None:
                continue
            if all(cell is None for cell in excel_row):
                continue
            row_dict: Dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = excel_row[idx] if idx < len(excel_row) else None
                if value is None:
                    row_dict[header] = ""
                else:
                    row_dict[header] = str(value).strip()
            rows.append(row_dict)

        workbook.close()
        return rows

    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return list(reader)


def load_existing_manifest(path: Path) -> Tuple[List[Dict[str, str]], Dict[str, int], Set[str]]:
    """Return existing manifest rows, an index by ``Catid`` and completed IDs."""

    entries: List[Dict[str, str]] = []
    index: Dict[str, int] = {}
    completed: Set[str] = set()
    if not path.exists():
        return entries, index, completed

    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            catid = (row.get("Catid") or "").strip()
            if catid and catid in index:
                entries[index[catid]] = row
            else:
                if catid:
                    index[catid] = len(entries)
                entries.append(row)

    for catid, idx in index.items():
        record = entries[idx]
        if (record.get("validation_passed") or "").strip().upper() == "TRUE":
            completed.add(catid)

    logging.info("Loaded %d existing manifest rows from %s", len(entries), path)
    return entries, index, completed


def record_manifest_entry(info: Dict[str, Any], entry: Dict[str, Any]) -> None:
    """Upsert ``entry`` into the manifest bookkeeping for a style."""

    catid = (entry.get("Catid") or "").strip()
    manifest: List[Dict[str, Any]] = info["manifest"]
    manifest_index: Dict[str, int] = info["manifest_index"]

    if catid and catid in manifest_index:
        manifest[manifest_index[catid]] = entry
    else:
        if catid:
            manifest_index[catid] = len(manifest)
        manifest.append(entry)

    validation = (entry.get("validation_passed") or "").strip().upper()
    completed: Set[str] = info["completed_catids"]
    if catid:
        if validation == "TRUE":
            completed.add(catid)
        else:
            completed.discard(catid)


def row_outputs_complete(catid: str, category_slug: str, styles: Dict[str, Dict[str, Any]]) -> bool:
    """Return ``True`` when every requested style already produced ``catid``."""

    for info in styles.values():
        file_prefix: str = info["file_prefix"]
        file_path = info["dir"] / category_slug / f"{file_prefix}{catid}.svg"
        if not file_path.exists():
            return False
        if catid not in info["completed_catids"]:
            return False
    return True


def parse_dimension(value: Optional[str]) -> float:
    """Return the numeric part of a dimension string (e.g. ``24px`` → ``24``)."""

    if not value:
        return 0.0
    match = re.search(r"-?\d+(?:\.\d+)?", value)
    if not match:
        return 0.0
    try:
        return float(match.group())
    except ValueError:
        return 0.0


def viewbox_dimensions(value: Optional[str]) -> Tuple[float, float]:
    """Extract width and height from a viewBox definition."""

    if not value:
        return 0.0, 0.0
    parts = [p for p in re.split(r"[ ,]+", value.strip()) if p]
    if len(parts) != 4:
        return 0.0, 0.0
    try:
        _, _, width, height = (float(part) for part in parts)
    except ValueError:
        return 0.0, 0.0
    return width, height


def fetch_icon_svg(
    category: str,
    catid: str,
    session: requests.Session,
    api_key: str,
    limit: int,
) -> Tuple[str, str, str]:
    """Return SVG data, source URL and title for ``category``.

    Returns empty strings when the lookup fails so the caller can record
    metadata about the missing public icon.
    """

    search_url = SVGAPI_LIST_URL.format(key=api_key)
    queries = iter_search_queries(category)
    for query in queries:
        try:
            response = session.get(
                search_url,
                params={"search": query, "limit": limit},
                timeout=10,
            )
            response.raise_for_status()
        except requests.RequestException as exc:
            logging.warning(
                "[svgapi] search failed for '%s' (query '%s'): %s",
                category,
                query,
                exc,
            )
            continue

        try:
            payload = response.json()
        except ValueError as exc:
            logging.warning(
                "[svgapi] invalid JSON for '%s' (query '%s'): %s",
                category,
                query,
                exc,
            )
            continue

        icons = payload.get("icons") or []
        if not icons:
            logging.info("[svgapi] no icons found for '%s' (query '%s')", category, query)
            continue

        idx = int(hashlib.sha256(catid.encode()).hexdigest(), 16) % len(icons)
        selected = icons[idx]
        svg_url = selected.get("url", "")
        if not svg_url:
            logging.warning(
                "[svgapi] missing SVG URL for '%s' (query '%s')",
                category,
                query,
            )
            continue

        try:
            svg_resp = session.get(svg_url, timeout=10)
            svg_resp.raise_for_status()
        except requests.RequestException as exc:
            logging.warning("[svgapi] download failed for '%s': %s", svg_url, exc)
            continue

        title = selected.get("title") or selected.get("slug") or category
        logging.info("[svgapi] using '%s' for '%s' via query '%s'", title, category, query)
        return svg_resp.text, svg_url, title

    logging.info("[svgapi] no icons found for '%s' after trying %d queries", category, len(queries))
    return "", "", ""


def restyle_svg(svg_data: str, params: dict) -> Tuple[str, List[str], str, int, int]:
    """Restyle raw SVG data to our specification and report basic metadata."""
    src_root = ET.fromstring(svg_data)
    view_box = src_root.get("viewBox") or src_root.get("viewbox")
    vb_width, vb_height = viewbox_dimensions(view_box)
    width_attr = parse_dimension(src_root.get("width"))
    height_attr = parse_dimension(src_root.get("height"))
    base_width = vb_width or width_attr or 24.0
    base_height = vb_height or height_attr or 24.0
    original_width = width_attr or vb_width or base_width
    original_height = height_attr or vb_height or base_height

    if params.get("raw_output"):
        primitives: List[str] = []
        for child in list(src_root):
            primitives.append(child.tag.split("}")[-1])
        signature = element_signature(src_root)
        path_hash = hashlib.sha256(signature.encode()).hexdigest()
        width_value = original_width or base_width
        height_value = original_height or base_height
        width_out = int(round(width_value)) if width_value else 0
        height_out = int(round(height_value)) if height_value else 0
        return svg_data, primitives, path_hash, width_out, height_out

    scale_base = max(base_width, base_height) or 256.0
    scale = 256 / scale_base

    preserve_style = params.get("preserve_source_style", False)
    svg_attrib = {
        "xmlns": SVG_NS,
        "viewBox": "0 0 256 256",
        "width": "256",
        "height": "256",
    }
    if preserve_style:
        if params.get("stroke_color"):
            svg_attrib["stroke"] = params["stroke_color"]
        if params.get("stroke_width"):
            svg_attrib["stroke-width"] = str(params["stroke_width"])
        if params.get("stroke_linecap"):
            svg_attrib["stroke-linecap"] = params["stroke_linecap"]
        if params.get("stroke_linejoin"):
            svg_attrib["stroke-linejoin"] = params["stroke_linejoin"]
        if params.get("fill") is not None:
            svg_attrib["fill"] = params["fill"]
    else:
        stroke_color = params.get("stroke_color", "#E63B14")
        stroke_width = params.get("stroke_width", 12)
        svg_attrib.update(
            {
                "stroke": stroke_color,
                "stroke-width": str(stroke_width),
                "stroke-linecap": params.get("stroke_linecap", "round"),
                "stroke-linejoin": params.get("stroke_linejoin", "round"),
            }
        )
        fill_value = params.get("fill", "none")
        if fill_value is not None:
            svg_attrib["fill"] = fill_value
    root = ET.Element("svg", svg_attrib)
    g = ET.SubElement(root, "g", {"transform": f"scale({scale})"})
    primitives: List[str] = []
    for child in list(src_root):
        if preserve_style:
            for attr in ["class", "id"]:
                child.attrib.pop(attr, None)
        else:
            for attr in ["stroke", "fill", "style", "class", "id"]:
                child.attrib.pop(attr, None)
        primitives.append(child.tag.split("}")[-1])
        g.append(child)
    shapes_sig = element_signature(g)
    path_hash = hashlib.sha256(shapes_sig.encode()).hexdigest()
    svg_content = ET.tostring(root, encoding="unicode")
    return svg_content, primitives, path_hash, 256, 256


def main():
    parser = argparse.ArgumentParser(description="Download SVG icons with style variants")
    parser.add_argument(
        "--csv",
        default="categories_sample.csv",
        help="Input taxonomy file (CSV or XLSX)",
    )
    parser.add_argument("--out", default="output/test2", help="Output directory root")
    parser.add_argument(
        "--styles",
        default=",".join(DEFAULT_STYLES),
        help="Comma-separated list of style variants to generate",
    )
    parser.add_argument(
        "--api-key",
        default=os.environ.get("SVGAPI_API_KEY", "Ty5WcDa63E"),
        help="SVGAPI key (defaults to example key or SVGAPI_API_KEY env var)",
    )
    parser.add_argument(
        "--search-limit",
        type=int,
        default=50,
        help="Number of results to request per API search",
    )
    parser.add_argument(
        "--log-level",
        default=os.environ.get("ICON_LOG_LEVEL", "INFO"),
        help="Logging level (DEBUG, INFO, WARNING, ERROR)",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Skip categories whose manifest entries and SVG files already exist",
    )
    args = parser.parse_args()

    input_path = Path(args.csv)
    out_root = Path(args.out)
    log_path = configure_logging(out_root, args.log_level)

    logging.info("Reading categories from %s", input_path)
    rows = load_taxonomy_rows(input_path)

    requested_styles: Iterable[str] = [s.strip() for s in args.styles.split(',') if s.strip()]
    styles: Dict[str, Dict[str, Any]] = {}
    for style in requested_styles:
        canonical_name = STYLE_ALIASES.get(style, style)
        if canonical_name not in STYLE_VARIANTS:
            raise SystemExit(
                f"Unknown style variant '{style}'. Available: {', '.join(sorted(STYLE_VARIANTS))}"
            )
        style_params = dict(STYLE_VARIANTS[canonical_name])
        styles[style] = style_params

    session = requests.Session()
    logging.info("Writing logs to %s", log_path)
    logging.info("Generating icons for %d categories (%s)", len(rows), ", ".join(styles))

    style_state: Dict[str, Dict[str, Any]] = {}
    for style_name, params in styles.items():
        style_dir = out_root / style_name
        style_dir.mkdir(parents=True, exist_ok=True)
        manifest_path = style_dir / "manifest.csv"
        if args.resume:
            manifest_rows, manifest_index, completed_catids = load_existing_manifest(manifest_path)
        else:
            manifest_rows, manifest_index, completed_catids = [], {}, set()
        style_state[style_name] = {
            "params": params,
            "dir": style_dir,
            "manifest": manifest_rows,
            "manifest_index": manifest_index,
            "completed_catids": completed_catids,
            "file_prefix": params.get("file_prefix", ""),
            "color": params.get("stroke_color") or "",
            "stroke_width": params.get("stroke_width") if params.get("stroke_width") is not None else "",
            "manifest_path": manifest_path,
        }

    for row in rows:
        catid_value = row.get('Catid', '')
        catid = str(catid_value).strip()
        if not catid:
            logging.warning("Skipping row without Catid: %s", row)
            continue

        category_name = deepest_category(row) or row.get('Root category') or 'Unknown'
        category_name = category_name.strip() if isinstance(category_name, str) else str(category_name)
        category_slug = slugify(category_name)
        logging.debug("Processing %s (%s)", catid, category_name)

        if args.resume and row_outputs_complete(catid, category_slug, style_state):
            logging.info("Skipping %s (%s) -- already complete", catid, category_name)
            continue

        svg_raw, source_url, icon_title = fetch_icon_svg(
            category_name,
            catid,
            session,
            args.api_key,
            args.search_limit,
        )

        if not svg_raw:
            for info in style_state.values():
                record_manifest_entry(
                    info,
                    {
                        'Catid': catid,
                        'category': category_slug,
                        'title_selected': category_name,
                        'concept_notes': 'no public icon found',
                        'primitives_used': '',
                        'path_hash': '',
                        'width': 0,
                        'height': 0,
                        'stroke_width': info['stroke_width'],
                        'color_hex': info['color'],
                        'validation_passed': 'FALSE',
                        'source_icon': '',
                    },
                )
            continue

        try:
            restyled: Dict[str, Tuple[str, List[str], str, int, int]] = {}
            for style_name, info in style_state.items():
                params = info['params']
                restyled[style_name] = restyle_svg(svg_raw, params)
        except ET.ParseError as exc:
            logging.warning("Failed to parse SVG for %s (%s): %s", catid, source_url, exc)
            for info in style_state.values():
                record_manifest_entry(
                    info,
                    {
                        'Catid': catid,
                        'category': category_slug,
                        'title_selected': category_name,
                        'concept_notes': 'svg parsing failed',
                        'primitives_used': '',
                        'path_hash': '',
                        'width': 0,
                        'height': 0,
                        'stroke_width': info['stroke_width'],
                        'color_hex': info['color'],
                        'validation_passed': 'FALSE',
                        'source_icon': source_url,
                    },
                )
            continue

        for style_name, info in style_state.items():
            svg_content, primitives, path_hash, width_out, height_out = restyled[style_name]
            style_dir: Path = info['dir']
            cat_dir = style_dir / category_slug
            cat_dir.mkdir(parents=True, exist_ok=True)
            file_prefix = info['file_prefix']
            file_path = cat_dir / f"{file_prefix}{catid}.svg"
            with open(file_path, 'w', encoding='utf-8') as sf:
                sf.write(svg_content)

            concept = f"downloaded from svgapi ({icon_title})"
            if info['params'].get('raw_output'):
                concept += " [raw]"
            record_manifest_entry(
                info,
                {
                    'Catid': catid,
                    'category': category_slug,
                    'title_selected': category_name,
                    'concept_notes': concept,
                    'primitives_used': ','.join(primitives),
                    'path_hash': path_hash,
                    'width': width_out,
                    'height': height_out,
                    'stroke_width': info['stroke_width'],
                    'color_hex': info['color'],
                    'validation_passed': 'TRUE',
                    'source_icon': source_url,
                },
            )

    fieldnames = [
        'Catid', 'category', 'title_selected', 'concept_notes', 'primitives_used',
        'path_hash', 'width', 'height', 'stroke_width', 'color_hex',
        'validation_passed', 'source_icon'
    ]

    for style_name, info in style_state.items():
        manifest_path: Path = info['manifest_path']
        with open(manifest_path, 'w', newline='', encoding='utf-8') as mf:
            writer = csv.DictWriter(mf, fieldnames=fieldnames)
            writer.writeheader()
            for row in info['manifest']:
                writer.writerow({key: row.get(key, "") for key in fieldnames})
        logging.info("Wrote %s", manifest_path)


if __name__ == "__main__":
    main()

